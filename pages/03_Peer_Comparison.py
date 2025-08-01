import streamlit as st
import pandas as pd
import matplotlib
import matplotlib.colors as mcolors
from pathlib import Path
import plotly.graph_objects as go
import numpy as np
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

## Page content. how it shows up on the side bar. how the page is laid out. wide in this case.
st.set_page_config(
    page_title="Peer Comparison",
    layout="wide",
)

## Set Page Title

st.title("Peer Comparison")

## Load the data. Cache so user only loads once upon use.

BASE_DIR = Path(__file__).resolve().parent.parent #file --> refers to where current py lives. parent parent goes up two levels

@st.cache_data
def load_all_excels():
    return (
        pd.read_excel(BASE_DIR/"transform_data.xlsx"),
        pd.read_excel(BASE_DIR/"raw_data.xlsx"),
        pd.read_excel(BASE_DIR/"coefficients_apr2024.xlsx"),
        pd.read_excel(BASE_DIR/"index_rating_scale.xlsx"),
        pd.read_excel(BASE_DIR/"index_variable_name.xlsx"),
        pd.read_excel(BASE_DIR/"index_country.xlsx"),
        pd.read_excel(BASE_DIR/"index_bbg_rating_live.xlsx", sheet_name="hard_code"),
    )
df_transform, df_raw, coeff_index, rating_index, variable_index, country_index, public_rating_index = load_all_excels()

## Define Loomis Colors for use later

LS_darkblue = "#1A3B73"
LS_lightblue = "#B6CEE4"
LS_faintblue = "#DAEEF3"
LS_darkgrey = "#91929B"
LS_lightgrey = "#E9E9EB"
LS_orange = "#EF7622"
excel_red = "#F8696B" 
excel_green = "#FFEB84"
excel_yellow = "#63BE7B"

## Load dictionaries for formatting and naming purposes. Highly efficient when using functions.

factors = ["wealth_factor",
           "size_factor",
           "growth_factor",
           "inflation_factor",
           "default_factor",
           "governance_factor",
           "fiscalperf_factor",
           "govdebt_factor",
           "extperf_factor",
           "reservebuffer_factor",
           "reservestatus_factor",
           ]

factors_dict = {
    "rating": "Avg Public Rating",
    "predicted_rating": "Model Rating",
    "wealth_factor":       "Wealth (5%)",
    "size_factor":         "Size (18%)",
    "growth_factor":       "Growth (2%)",
    "inflation_factor":    "Inflation (3%)",
    "default_factor":      "Default History (9%)",
    "governance_factor":   "Governance (32%)",
    "fiscalperf_factor":   "Fiscal Performance (7%)",
    "govdebt_factor":      "Government Debt (10%)",
    "extperf_factor":      "External Performance (5%)",
    "reservebuffer_factor":"FX Reserves (4%)",
    "reservestatus_factor":"Reserve Currency Status (5%)"
}

variable_dict = {
    "ngdp_pc":       "Nominal GDP per capita (US$)",
    "ngdp":         "Nominal GDP (bil US$)",
    "growth_avg":       "Avg 10Yr GDP Growth t-5 to t+4 (%)",
    "inf_avg":    "Average 10Yr Inflation t-5 to t+4 (%)",
    "default_hist":      "Default History Dummy (1=Yes, 0=No)",
    "default_decay":   "Default Decay (1 at incidence)",
    "voice_acct":   "Voice and Accountability (Z-score)",
    "pol_stab":      "Political Stability (Z-score)",
    "gov_eff":      "Government Effectiveness (Z-score)",
    "reg_qual":"Regulatory Quality (Z-score)",
    "rule_law":"Rule of Law (Z-score)",
    "cont_corrupt": "Control of Corruption (Z-score)",
    "fb_avg": "Avg 10Yr Fiscal Balance t-5 to t+4 (% of GDP)",
    "gov_rev_gdp": "Government Revenue (% of GDP)",
    "ir_rev": "Interest Payment (% of Revenue)",
    "gov_debt_gdp": "Government Debt (% of GDP)",
    "cab_avg": "Avg 10Yr Current Account Balance t-5 to t+4 (% of GDP)",
    "reserve_gdp": "FX Reserves (% of GDP)",
    "import_cover": "FX Reserves (months of imports)",
    "reserve_fx": "Reserve Currency Status (1 = Yes, 0 = No)"

}

# map each short_var to a Python format‐string for its value
format_map = {
    "ngdp_pc":"${value:,.0f}",
    "ngdp":"${value:,.1f} bil",
    "growth_avg": "{value:.1f}%",
    "inf_avg":    "{value:.1f}%",
    "default_hist": "{value:.0f}",
    "default_decay": "{value:.2f}",
    "voice_acct": "{value:.2f}",
    "pol_stab": "{value:.2f}",
    "gov_eff": "{value:.2f}",
    "reg_qual":"{value:.2f}",
    "rule_law":"{value:.2f}",
    "cont_corrupt": "{value:.2f}",
    "fb_avg": "{value:.1f}%",
    "gov_rev_gdp": "{value:.1f}%",
    "ir_rev": "{value:.1f}%",
    "gov_debt_gdp": "{value:.1f}%",
    "cab_avg": "{value:.1f}%",
    "reserve_gdp": "{value:.1f}%",
    "import_cover": "{value:.1f}",
    "reserve_fx": "{value:.0f}"}

## Toggle rating buckets. We need this to help users get an idea of which countries are in a rating bucket

# Make copies of df_transform and df_raw. Add rounded rating col to help with fitering.

rounded_transform = df_transform['rating'].round()
df_transform.insert(3,'round_rating',rounded_transform)

rounded_raw = df_raw['rating'].round()
df_raw.insert(3,'round_rating',rounded_raw)

rating_ranges = {
    "AAA": (22, 22),
    "AA":  (19, 21),
    "A":   (16, 18),
    "BBB": (13, 15),
    "BB":  (10, 12),
    "B":  (7, 9),
    "CCC to C": (2,6),
    "D": (0,1),
    "ALL": (0, 22),
    "IG": (13, 22),
    "HY": (1, 12)}

#zip pairs the two columns row by row to help make into a dict
rating_dict = dict(zip(rating_index['Numeric'], rating_index['Credit Rating'])) 

## Set up select boxes and fitlered dfs

# Limit the width of the select boxes
st.markdown("""
<style>
/* 🔹 Limit the max width of selectboxes */
div[data-baseweb="select"] {
    max-width: 300px !important;
}
</style>
""", unsafe_allow_html=True)

# Drop down to select years

available_year = df_transform['year'].unique()
selected_year = st.selectbox("Select Year", sorted(available_year, reverse=True))

# Subset our data frame to only include the selected time frame (selected_year) 

df_transform_filter = df_transform[df_transform["year"] == selected_year]
df_raw_filter = df_raw[df_raw["year"] == selected_year]

# Drop down select rating ranges
# Note we don't subset the df further over here.. this is just to produce a list of plausible peers for our users

#user selects key from rating_ranges dictionary
selected_bucket = st.selectbox("Show me all countries that are rated...", list(rating_ranges.keys())) 
#Lookup up lower and upper bound based on rating category selected (tuple unpacking)
low, high = rating_ranges[selected_bucket]

# filter the DataFrame
matches = (
    df_transform_filter['round_rating']
      .between(low, high)
) #generates a boolean list. true if round_rating meets the criteria. false otherwise

countries = df_transform_filter.loc[matches, 'name'].tolist()
#using that boolean list, we pull out only the names that conform to true
#we transform this to a list called countries

# format into a single line
if countries:
    line = ", ".join(countries)
else:
    line = "(no countries found)"

# Print it out in streamlit
st.write(line)

# Multiselct to allow user to select up to five countries to compare

# set it such that this select box gets to be wider

#this controls width of the box
st.markdown("""
<style>
/* override just that multiselect */
div[data-testid="stMultiSelect-country_compare"] 
  div[data-baseweb="select"] {
    max-width: 600px !important;
}
</style>
""", unsafe_allow_html=True)

peer_list = sorted(df_transform_filter["name"].unique().tolist())

peers = st.multiselect(
    "Choose up to FIVE countries to compare",
    options = peer_list,
    #default = peer_list[:5],
    max_selections = 5,
    key = "country_compare"
)

####----Start Building the Factor Level Table here now that the pre-requisites are set----####

# 2) Assume `peers` is your list of five country names from the multiselect
#    and df_transform_filter is already filtered to the right year.
#    e.g. peers = st.session_state.country_compare
#          df = df_transform_filter[df_transform_filter["name"].isin(peers)]

# 3) Build a row for each factor

rows = []
for short_var, display_name in factors_dict.items():
    row = {"Variable": display_name}
    for country in peers:
        # Pull the raw value
        raw = (
            df_transform_filter
            .loc[df_transform_filter["name"] == country, short_var]
            .iloc[0]
        )

        if short_var in ["rating", "predicted_rating"]:
            # Round to nearest integer and map to letter (fallback to blank)
            #letter = rating_dict.get(round(raw), "") --> revive later if need letter rating
            row[country] = raw
        else:
            # Round everything else to 2 decimal places
            row[country] = raw

    rows.append(row)

# 4) Assemble into a DataFrame
heatmap_df = pd.DataFrame(rows)

#— Remove the “Variable” header by making it the index —
heatmap_df = heatmap_df.set_index("Variable")
#heatmap_df.index.name = None   # ← removes the “Variable” label in the corner

# “Reset” the index so “Variable” becomes a real column again,
# and the DataFrame’s index is the default integers (0, 1, 2, …).
heatmap_df = heatmap_df.reset_index()

# 5) Now we use the .style property to decorate our df and then use st.write to render it so it looks the way we want

# Identify which row labels need letter formatting

letter_rows = ["Avg Public Rating", "Model Rating"]
numeric_rows = [r for r in heatmap_df["Variable"] if r not in letter_rows]

# Customize how you want your color shading map to look

low_color  = "#F8696B"   # excel conditional red
mid_color = "#FFEB84" # excel conditional yellow
high_color = "#63BE7B"   # excel conditional green

#Create a LinearSegmentedColormap that goes from red → green:

excel_cmap = mcolors.LinearSegmentedColormap.from_list(
    "excel_r_y_g",
    [low_color, mid_color, high_color]
)


#create Helper function to colour cells in the “Variable” column

def style_variable_cell(row: pd.Series) -> pd.Series:
    """
    row.name is the integer index (0, 1, 2, …).
    We return a Series of CSS strings, one per column, for this single row:
      • If the column is "Variable" and row.name < 2 → LS_orange bg, black text, bold.
      • If the column is "Variable" and row.name >= 2 → LS_lightblue bg, black text, bold.
      • For all other columns, return "" (no styling).
    """
    out = pd.Series("", index=row.index)
    idx = row.name

    if idx < 2:
        out["Variable"] = (
            f"background-color: {LS_orange}; "
            "color: black; "
            "font-weight: bold;"
        )
    else:
        out["Variable"] = (
            f"background-color: {LS_faintblue}; "
            "color: black; "
            "font-weight: bold;"
        )
    return out


#def color_label_cell(val):
#    return "background-color: #1A3B73; color: white;"

# Create style object that adds various style elements to the df

styler = (
    heatmap_df.style
        # (a) Put a gradient on numeric rows only:
        .background_gradient(
            cmap=excel_cmap,
            subset=pd.IndexSlice[
                heatmap_df["Variable"].isin(numeric_rows),  # select rows by boolean mask
                heatmap_df.columns.difference(["Variable"]) # all columns except “Variable”
            ],axis=1 #apply condtional format across rows
        )
        # (b) For letter_rows, convert the float into the letter:
        .format(
            lambda v: rating_dict.get(int(round(v)), ""),
            subset=pd.IndexSlice[
                heatmap_df["Variable"].isin(letter_rows),  # only letter rows
                heatmap_df.columns.difference(["Variable"])
            ]
        )
        # (c) For numeric_rows, display exactly 2 decimals:
        .format(
            "{:.2f}",
            subset=pd.IndexSlice[
                heatmap_df["Variable"].isin(numeric_rows),
                heatmap_df.columns.difference(["Variable"])
            ]
        )
        # apply style_variable_cell to our whole df. this function mainly defines how i want the variable column to look
        .apply(style_variable_cell, axis=1)
        # (e) Hide Streamlit’s default integer index (pandas ≥ 1.4.0)
        #.hide_index()
)

# Render in Streamlit

#st.table(styler) if you want the simple versionw without the interactivity
st.subheader("Rating Factor Heat Map (Z-scores)")
st.write(styler)

# Create excel export button below

def generate_export_short(df: pd.DataFrame) -> BytesIO:
    wb = Workbook()
    ws = wb.active

    # 1a) Write the header row (your country names) into row 1
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # 1b) Write the DataFrame’s values from row 2 onward
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # — 2)Recompute dimensions
    max_row = ws.max_row
    max_col = ws.max_column

    # — 3) Auto-fit column A width —
    colA = get_column_letter(1)
    max_w = max(
        len(str(ws[f"{colA}{r}"].value or ""))
        for r in range(1, max_row + 1)
    )
    ws.column_dimensions[colA].width = max_w + 1
    
    # — 4) Delete contents of A1 —
    ws["A1"].value = None

    # — 5) Shade & bold A2–A14 —
    orange = PatternFill("solid", fgColor="FFEF7622")
    light_tint = PatternFill("solid", fgColor="FFDAEEF3")
    bold = Font(bold=True)
    for r in range(2, 15):
        cell = ws.cell(row=r, column=1)
        cell.font = bold
        if r in (2, 3):
            cell.fill = orange
        else:
            cell.fill = light_tint

    # — 6) Header row B1–F1 styling if populated —
    header_fill = PatternFill("solid", fgColor="FF1A3B73")
    white_font = Font(color="FFFFFFFF", bold=True)
    for col_idx in range(2, 7):
        cell = ws.cell(row=1, column=col_idx)
        if cell.value not in (None, ""):
            cell.alignment = Alignment(wrapText=True)
            cell.font      = white_font
            cell.fill      = header_fill
    
    # — 6b) just force col B1 to F1 to be width 10.0
    for col in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 12.0

    # — 7) Numeric formatting B4–F14 & per-row 3-color scale —
    color_rule = ColorScaleRule(
        start_type='min',  start_color='FFF8696B',
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max',   end_color='FF63BE7B'
        )

    for r in range(4, 15):
        # 1) format numbers
        for col_idx in range(2, 7):
            cell = ws.cell(row=r, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"
        # 2) apply a single color-scale rule to the entire row range B…F
        rng = f"B{r}:F{r}" #That line is a Python f-string that builds the Excel range address for columns B through F on row r.
        ws.conditional_formatting.add(rng, color_rule) 
               
    # — 8) Map B2–F3 from numeric → letter rating via rating_dict —
    for r in (2, 3):
        for col_idx in range(2, 7):
            cell = ws.cell(row=r, column=col_idx)
            if isinstance(cell.value, (int, float)):
                num = round(cell.value)
                letter = rating_dict.get(num, "")
                cell.value = letter
                cell.alignment = Alignment(horizontal="center")

    # — 9) Draw a full border around A1:F14 —
    thin = Side(style="thin")
    def mk_border(top=False, bottom=False, left=False, right=False):
        return Border(
            top    = thin if top    else Side(style=None),
            bottom = thin if bottom else Side(style=None),
            left   = thin if left   else Side(style=None),
            right  = thin if right  else Side(style=None),
        )
    for r in range(1, 15):
        for c in range(1, 7):
            if r in (1, 14) or c in (1, 6):
                ws.cell(row=r, column=c).border = mk_border(
                    top    = (r == 1),
                    bottom = (r == 14),
                    left   = (c == 1),
                    right  = (c == 6)
                )

    # — 10) Save to in-memory buffer —
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

excel_data_short = generate_export_short(heatmap_df)

export_col_short, blank_col_1, blank_col_2 = st.columns([2, 2, 6])

with export_col_short:
    st.download_button(
    label="📥 Export to Excel (Formatted)",
    key="short_excel",
    data=excel_data_short,
    file_name="factor_heatmap.xlsx",
    mime="application/vnd.openxmlformats-officedocument-spreadsheetml.sheet"
)

####----Start Building the Constituent Level HeatMap here now that the pre-requisites are set----####

## First we form the df we want, using the same method as above

rows_long = []

# Get "rating" and "predicted_rating" from transform_df_filter

desired_vars = ["rating", "predicted_rating"]

for short_var in desired_vars:
    # Look up the display name from factors_dict
    display_name = factors_dict[short_var]
    row_short = {"Variable": display_name}
    
    for country in peers:
        # Pull the raw value for this country & variable
        raw = (
            df_transform_filter
            .loc[df_transform_filter["name"] == country, short_var]
            .iloc[0]
        )

        # If it’s one of the two letter‐rating fields, leave as is; otherwise you could format.
        # (Here we just assign raw for both since they are both “rating” fields.)
        row_short[country] = raw

    rows_long.append(row_short)

# Get the other raw variables from df_raw_filter

for short_var, display_name in variable_dict.items():
    row_long = {"Variable": display_name}
    for country in peers:
        # Pull the raw value
        raw = (
            df_raw_filter
            .loc[df_raw_filter["name"] == country, short_var]
            .iloc[0]
        )

        row_long[country] = raw

    rows_long.append(row_long)

#Assemble into a DataFrame
heatmap_df_long = pd.DataFrame(rows_long)

#Render the output to check if we made our df properly
st.subheader("Constituent Variable Heat Map (raw numerical values)")

#now we do the difficult step of slicing and recombining our df so that we can get row headers!

#first we slice up our df into where we want to slot
df1 = heatmap_df_long.iloc[:2]
df2 = heatmap_df_long.iloc[2:3]
df3 = heatmap_df_long.iloc[3:4]
df4 = heatmap_df_long.iloc[4:5]
df5 = heatmap_df_long.iloc[5:6]
df6 = heatmap_df_long.iloc[6:8]
df7 = heatmap_df_long.iloc[8:14]
df8 = heatmap_df_long.iloc[14:17]
df9 = heatmap_df_long.iloc[17:18]
df10 = heatmap_df_long.iloc[18:19]
df11 = heatmap_df_long.iloc[19:21]
df12 = heatmap_df_long.iloc[21:22]

#next we create blank rows for the headers we want

# Get a list of all columns except "Variable"
country_cols = [col for col in heatmap_df_long.columns if col != "Variable"]
#We know we always have exactly one column named "Variable", and all the others (1 or more) are the country names chosen by the user.

wealth_dict = {"Variable": "Wealth (5%)"}
for c in country_cols:
    wealth_dict[c] = 0.0    # ← or use "" if you dislike 0.0. looping over dynamic list of countries
wealth_row = pd.DataFrame([wealth_dict]) #turn dict into a df

size_dict = {"Variable": "Size (18%)"}
for c in country_cols:
    size_dict[c] = 0.0    # ← or use "" if you dislike 0.0. looping over dynamic list of countries
size_row = pd.DataFrame([size_dict]) #turn dict into a df

growth_dict = {"Variable": "Growth (2%)"}
for c in country_cols:
    growth_dict[c] = 0.0    # ← or use "" if you dislike 0.0. looping over dynamic list of countries
growth_row = pd.DataFrame([growth_dict]) #turn dict into a df

inf_dict = {"Variable": "Inflation (3%)"}
for c in country_cols:
    inf_dict[c] = 0.0    # ← or use "" if you dislike 0.0. looping over dynamic list of countries
inf_row = pd.DataFrame([inf_dict]) #turn dict into a df

default_dict = {"Variable": "Default History (9%)"}
for c in country_cols:
    default_dict[c] = 0.0    # ← or use "" if you dislike 0.0. looping over dynamic list of countries
default_row = pd.DataFrame([default_dict]) #turn dict into a df

gov_dict = {"Variable": "Governance (32%)"}
for c in country_cols:
    gov_dict[c] = 0.0    # ← or use "" if you dislike 0.0. looping over dynamic list of countries
gov_row = pd.DataFrame([gov_dict]) #turn dict into a df

fiscal_dict = {"Variable": "Fiscal Performance (7%)"}
for c in country_cols:
    fiscal_dict[c] = 0.0    # ← or use "" if you dislike 0.0. looping over dynamic list of countries
fiscal_row = pd.DataFrame([fiscal_dict]) #turn dict into a df

debt_dict = {"Variable": "Government Debt (10%)"}
for c in country_cols:
    debt_dict[c] = 0.0    # ← or use "" if you dislike 0.0. looping over dynamic list of countries
debt_row = pd.DataFrame([debt_dict]) #turn dict into a df

ext_dict = {"Variable": "External Performance (5%)"}
for c in country_cols:
    ext_dict[c] = 0.0    # ← or use "" if you dislike 0.0. looping over dynamic list of countries
ext_row = pd.DataFrame([ext_dict]) #turn dict into a df

fx_dict = {"Variable": "FX Reserves (4%)"}
for c in country_cols:
    fx_dict[c] = 0.0    # ← or use "" if you dislike 0.0. looping over dynamic list of countries
fx_row = pd.DataFrame([fx_dict]) #turn dict into a df

reserve_dict = {"Variable": "Reserve Currency Status (5%)"}
for c in country_cols:
    reserve_dict[c] = 0.0    # ← or use "" if you dislike 0.0. looping over dynamic list of countries
reserve_row = pd.DataFrame([reserve_dict]) #turn dict into a df

heatmap_df_long = pd.concat(
    [
      df1, wealth_row,
      df2, size_row,
      df3, growth_row,
      df4, inf_row,
      df5, default_row,
      df6, gov_row,
      df7, fiscal_row,
      df8, debt_row,
      df9, ext_row,
      df10, fx_row,
      df11, reserve_row,
      df12
    ],
    ignore_index=True
)

# now time to get the styler in place so we can make the table look the way we want!

# Identify which row labels need letter formatting

letter_rows_long = ["Avg Public Rating",
                    "Model Rating",
                    "Wealth (5%)",
                    "Size (18%)",
                    "Growth (2%)",
                    "Inflation (3%)",
                    "Default History (9%)",
                    "Governance (32%)",
                    "Fiscal Performance (7%)",
                    "Government Debt (10%)",
                    "External Performance (5%)",
                    "FX Reserves (4%)",
                    "Reserve Currency Status (5%)"
                    ]

numeric_rows_long = [r for r in heatmap_df_long["Variable"] if r not in letter_rows_long]

header_rows_long = ["Wealth (5%)",
                    "Size (18%)",
                    "Growth (2%)",
                    "Inflation (3%)",
                    "Default History (9%)",
                    "Governance (32%)",
                    "Fiscal Performance (7%)",
                    "Government Debt (10%)",
                    "External Performance (5%)",
                    "FX Reserves (4%)",
                    "Reserve Currency Status (5%)"
                    ]

#use this function to hide zeros and make them appear as "". used for header columns
def hide_zeros(v):
    # If it’s exactly 0 or 0.0, return an empty string
    if pd.isna(v):
        return ""       # (optional) leave NA values blank, too
    if float(v) == 0.0:
        return ""
    # Otherwise format to two decimals
    return f"{v:.2f}"

one_dp_comma = ["Nominal GDP per capita (US$)",
                "Nominal GDP (bil US$)"]

one_dp = ["Avg 10Yr GDP Growth t-5 to t+4 (%)",
          "Average 10Yr Inflation t-5 to t+4 (%)",
          "Default Decay (1 at incidence)",
          "Avg 10Yr Fiscal Balance t-5 to t+4 (% of GDP)",
          "Government Revenue (% of GDP)",
          "Interest Payment (% of Revenue)",
          "Government Debt (% of GDP)",
          "Avg 10Yr Current Account Balance t-5 to t+4 (% of GDP)",
          "FX Reserves (% of GDP)",
          "FX Reserves (months of imports)",
          "Reserve Currency Status (1 = Yes, 0 = No)"]

two_dp = [
        "Voice and Accountability (Z-score)",
        "Political Stability (Z-score)",
        "Government Effectiveness (Z-score)",
        "Regulatory Quality (Z-score)",
        "Rule of Law (Z-score)",
        "Control of Corruption (Z-score)",]

dummy_dp = ["Default History Dummy (1=Yes, 0=No)",
            "Reserve Currency Status (1 = Yes, 0 = No)"]

def style_variable_cell_long(row: pd.Series) -> pd.Series:
    """
    row.name is the integer index (0, 1, 2, …).
    row["Variable"] is the label, e.g. "Avg Public Rating", "Wealth (5%)", etc.

    We return a Series of CSS strings—one per column in this row—according to:
      • If row["Variable"] is "Avg Public Rating" or "Model Rating": 
            shade every cell in this row LS_orange, black bold text.
      • If row["Variable"] is one of the other specified variables (Wealth, Size, etc.):
            shade every cell in this row LS_faintblue, black bold text.
      • Else: leave all cells in this row unstyled ("").
    """
    out = pd.Series("", index=row.index)
    var = row["Variable"]

    orange_set = {"Avg Public Rating", "Model Rating"}
    blue_set = {
        "Wealth (5%)",
        "Size (18%)",
        "Growth (2%)",
        "Inflation (3%)",
        "Default History (9%)",
        "Governance (32%)",
        "Fiscal Performance (7%)",
        "Government Debt (10%)",
        "External Performance (5%)",
        "FX Reserves (4%)",
        "Reserve Currency Status (5%)"
    }

    if var in orange_set:
        css = f"background-color: {LS_orange}; color: black; font-weight: bold;"
        out["Variable"] = css
    elif var in blue_set:
        css = f"background-color: {LS_faintblue}; color: black; font-weight: bold;"
        out[:] = css
    # else: leave out[:] as "" so no styling

    return out

styler_long = (
    heatmap_df_long.style
        # (a) Put a gradient on numeric rows only:
        .background_gradient(
            cmap=excel_cmap, #note that excel_cmap is defined above already when making the short table
            subset=pd.IndexSlice[
                heatmap_df_long["Variable"].isin(numeric_rows_long),  # which rows to style
                heatmap_df_long.columns.difference(["Variable"]) # which columns to style
            ],axis=1
        )
        # (b) For letter_rows, convert the float into the letter:
        .format(
            lambda v: rating_dict.get(int(round(v)), ""),
            subset=pd.IndexSlice[
                heatmap_df_long["Variable"].isin(["Avg Public Rating", "Model Rating"]),  # which rows to style
                heatmap_df_long.columns.difference(["Variable"]) # which columns to style
            ]
        )
        # hide the zeros in the header columns
        .format(
            hide_zeros,
            subset=pd.IndexSlice[
                heatmap_df_long["Variable"].isin(header_rows_long),
                heatmap_df_long.columns.difference(["Variable"])
            ]
        )

        # 1dp with , for thousands
        .format(
            "{:,.1f}",
            subset=pd.IndexSlice[
                heatmap_df_long["Variable"].isin(one_dp_comma),
                heatmap_df_long.columns.difference(["Variable"])
            ]
        )

        # 1dp format
        .format(
            "{:.1f}",
            subset=pd.IndexSlice[
                heatmap_df_long["Variable"].isin(one_dp),
                heatmap_df_long.columns.difference(["Variable"])
            ]
        )

        # 2dp format
        .format(
            "{:.2f}",
            subset=pd.IndexSlice[
                heatmap_df_long["Variable"].isin(two_dp),
                heatmap_df_long.columns.difference(["Variable"])
            ]
        )

        # dummy whole number format
        .format(
            "{:.0f}",
            subset=pd.IndexSlice[
                heatmap_df_long["Variable"].isin(dummy_dp),
                heatmap_df_long.columns.difference(["Variable"])
            ]
        )

        # apply style_variable_cell to our whole df. this function mainly defines how i want the variable column to look
        .apply(style_variable_cell_long, axis=1)
        # (e) Hide Streamlit’s default integer index (pandas ≥ 1.4.0)
        #.hide_index()
)

# LEts render the table nicely

st.write(styler_long)

# Create excel export button below

def generate_export_long(df: pd.DataFrame) -> BytesIO:
    wb = Workbook()
    ws = wb.active

    # 1a) Write the header row (your column names) into row 1
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # 1b) Write the DataFrame’s values from row 2 onward (skipping the index)
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    
    # 2) Recompute sheet dimensions for later loops
    max_row = ws.max_row
    max_col = ws.max_column

    # 3) Auto-fit column A
    colA = get_column_letter(1)
    max_w = max(len(str(ws[f"{colA}{r}"].value or "")) for r in range(1, ws.max_row+1))
    ws.column_dimensions[colA].width = max_w + 2

    # 4) Clear A1
    ws["A1"].value = None

    # Styles & helpers
    orange_fill = PatternFill("solid", fgColor="FFEF7622")
    tint_fill   = PatternFill("solid", fgColor="FFDAEEF3")
    header_fill = PatternFill("solid", fgColor="FF1A3B73")
    bold_font   = Font(bold=True)
    white_bold  = Font(bold=True, color="FFFFFFFF")
    wrap        = Alignment(wrapText=True)
    center      = Alignment(horizontal="center")
    thin        = Side(style="thin")
    def mk_border(top=False, bottom=False, left=False, right=False):
        return Border(
            top    = thin if top    else Side(style=None),
            bottom = thin if bottom else Side(style=None),
            left   = thin if left   else Side(style=None),
            right  = thin if right  else Side(style=None),
        )

    # 5) Bold A2–A34
    for r in range(2, 35):
        ws.cell(row=r, column=1).font = bold_font

    # 6) Shade A2–A3 orange
    for r in (2,3):
        ws.cell(row=r, column=1).fill = orange_fill

    # 7) For each of these rows, clear B–F and shade A–F light tint
    for r in (4,6,8,10,12,15,22,26,28,30,33):
        for c in range(2,7):
            ws.cell(row=r, column=c).value = None
        for c in range(1,7):
            ws.cell(row=r, column=c).fill = tint_fill

    # 8) Header row B1–F1: wrap, bold white font, dark-blue fill
    for c in range(2,7):
        cell = ws.cell(row=1, column=c)
        if cell.value not in (None, ""):
            cell.alignment = wrap
            cell.font      = white_bold
            cell.fill      = header_fill

    # 9) Force columns B–F to width 12
    for c in range(2,7):
        ws.column_dimensions[get_column_letter(c)].width = 12.0

    # 10) Number‐format each specific row range
    # B5–F5: whole w/ comma
    for c in range(2,7):
        cell = ws.cell(row=5, column=c)
        if isinstance(cell.value, (int,float)):
            cell.number_format = "#,##0"
    # B7–F7: 1 dp w/ comma
    for c in range(2,7):
        cell = ws.cell(row=7, column=c)
        if isinstance(cell.value, (int,float)):
            cell.number_format = "#,##0.0"
    # One‐decimal rows:
    one_dp = [9,11,16,17,18,19,20,21,23,24,25,27,29,31,32]
    for r in one_dp:
        for c in range(2,7):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int,float)):
                cell.number_format = "0.0"
    # B13–F13 & B34–F34: whole
    for r in (13,34):
        for c in range(2,7):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int,float)):
                cell.number_format = "#,##0"
    # B14–F14: 2 dp but show "0" if zero
    for c in range(2,7):
        cell = ws.cell(row=14, column=c)
        if isinstance(cell.value, (int,float)):
            if cell.value == 0:
                cell.number_format = "0"
            else:
                cell.number_format = "0.00"

    # 11) Conditional 3-color scales
    high_good = ColorScaleRule(
        start_type='min',  start_color='FFF8696B',
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max',    end_color='FF63BE7B'
    )
    low_good = ColorScaleRule(
        start_type='min',  start_color='FF63BE7B',
        mid_type='percentile', mid_value=50, mid_color='FFEB84',
        end_type='max',    end_color='FFF8696B'
    )
    high_rows = [5,7,9,16,17,18,19,20,21,23,24,29,31,32,34]
    low_rows  = [11,13,14,25,27]
    for r in high_rows:
        rng = f"B{r}:F{r}"
        ws.conditional_formatting.add(rng, high_good)
    for r in low_rows:
        rng = f"B{r}:F{r}"
        ws.conditional_formatting.add(rng, low_good)

    # 12) Map numeric B2–F3 → letter via rating_dict
    for r in (2,3):
        for c in range(2,7):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if isinstance(v, (int,float)):
                letter = rating_dict.get(round(v), "")
                cell.value     = letter
                cell.alignment = center

    # 13) Draw a thin border around A1:F34
    for r in range(1, 35):
        for c in range(1, 7):
            if r in (1,34) or c in (1,6):
                ws.cell(row=r, column=c).border = mk_border(
                    top    = (r == 1),
                    bottom = (r == 34),
                    left   = (c == 1),
                    right  = (c == 6)
                )

    # 14) Save to BytesIO
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

excel_data_long = generate_export_long(heatmap_df_long)

export_col_long, blank_col_1, blank_col_2 = st.columns([2, 2, 6])

with export_col_long:
    st.download_button(
    label="📥 Export to Excel (Formatted)",
    key="long_excel",
    data=excel_data_long,
    file_name="variable_heatmap.xlsx",
    mime="application/vnd.openxmlformats-officedocument-spreadsheetml.sheet"
)