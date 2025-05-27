import streamlit as st
import pandas as pd
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode, ColumnsAutoSizeMode
import os #--> helps to save user edits on to pc
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import gspread
from google.oauth2.service_account import Credentials
from gsheets_utils import load_override_from_gsheet, save_override_to_gsheet

# Page setup. (must be your very first Streamlit call)

st.set_page_config(
    page_title="LS Sovereign Credit Rating Model",
    layout="wide",
)
# Page content

st.title("LS Sovereign Credit Rating Model")

# Load data
# Caches the data so that it only loads once for the users. since the dfs don't change much.
# this reduces the loading time

#if st.button("🔄 Refresh data from source"): #--> to use in future if i run the paid version!
    #st.cache_data.clear()        # clear ALL @st.cache_data caches
    #st.rerun()      # immediately rerun the script

@st.cache_data
def load_all_excels():
    return (
        pd.read_excel("transform_data.xlsx"),
        pd.read_excel("raw_data.xlsx"),
        pd.read_excel("coefficients_apr2024.xlsx"),
        pd.read_excel("index_rating_scale.xlsx"),
        pd.read_excel("index_variable_name.xlsx"),
        pd.read_excel("index_country.xlsx"),
        pd.read_excel("index_bbg_rating_live.xlsx", sheet_name="hard_code"),
    )
df_transform, df_raw, coeff_index, rating_index, variable_index, country_index, public_rating_index = load_all_excels()


#Inject the width-limiting CSS before your selectbox calls
#else they appeared to be too wide!

st.markdown("""
<style>
/* 🔹 Limit the max width of selectboxes */
div[data-baseweb="select"] {
    max-width: 300px !important;
}
</style>
""", unsafe_allow_html=True)

# Dropdown to select Country

country_name = df_transform['name'].unique()
selected_name = st.selectbox("Select Country", sorted(country_name))

# Dropdown to select years

filtered_year = df_transform[df_transform['name'] == selected_name]['year'].unique()
selected_year = st.selectbox("Select Year", sorted(filtered_year, reverse=True))

# Select Row based on year and country

selected_row = df_transform[(df_transform['name'] == selected_name) & (df_transform['year'] == selected_year)]

# Select factors and ratings from transform_df to show in table

select_cols = ["wealth_factor",
                "size_factor",
                "growth_factor",
                "inflation_factor",
                "default_factor",
                "governance_factor",
                "fiscalperf_factor",
                "govdebt_factor",
                "extperf_factor",
                "reservebuffer_factor",
                "reservestatus_factor"
                ] 

short_table_df = selected_row[select_cols].T.reset_index() #T is transpose. Reset Index just prevents variable name from being index
short_table_df.columns = ['short_name', 'Z-score Value'] #rename columns

#insert row for constant factor / regression intercept

const_row = pd.DataFrame([{
    'short_name' : 'const',
    'Z-score Value' : float(1.0)

}])

short_table_df = pd.concat([const_row, short_table_df], ignore_index=True)

#merge regression coefficients into the table

## rename first col to 'short_name' --> which we use to stitch everything together
coeff_index_short = coeff_index.copy()
coeff_index_short = coeff_index_short.rename(columns={'Unnamed: 0': 'short_name'})

## merge on 'short_name' into short_table_df
short_table_df = pd.merge(short_table_df, coeff_index_short, on = 'short_name', how='left')

#merge long names into the table

## rename first col to 'short_name' which we use to stitch everything together
variable_index_short = variable_index.copy()
variable_index_short = variable_index.rename(columns={'Unnamed: 0': 'short_name'})

## merge on 'short_name' into short_table_df
short_table_df = pd.merge(short_table_df, variable_index_short, on = 'short_name', how='left')

#drop description column

short_table_df = short_table_df.drop(columns=['description'])

#rearrange columns

short_table_df = short_table_df[['short_name','long_name','coefficient','Z-score Value']]

#add in rating column which is beta * X
short_table_df['Rating (notches)'] = short_table_df['coefficient'] * short_table_df['Z-score Value']

# Add final row at the bottom which is the model predicted rating (numeric)

## Step 1: Calculate the total
model_rating = short_table_df['Rating (notches)'].sum()

## Step 2: Create the total row
total_row = pd.DataFrame([{
    'short_name': 'predicted_rating',
    'long_name': 'Model Rating',
    'coefficient': '',
    'Z-score Value': '',
    'Rating (notches)': model_rating
}])

## Step 3: Append it to the bottom
short_table_df = pd.concat([short_table_df, total_row], ignore_index=True)

# Manually Insert Section Headers

## Build section header rows
economy_pillar = pd.DataFrame([{
    'short_name': 'eco_header',
    'long_name': 'REAL ECONOMY PILLAR (25%)',
    'coefficient': '',
    'Z-score Value': '',
    'Rating (notches)': ''
}])

institutions_pillar = pd.DataFrame([{
    'short_name': 'insti_header',
    'long_name': 'MONETARY & INSTITUTIONS PILLAR (44%)',
    'coefficient': '',
    'Z-score Value': '',
    'Rating (notches)': ''
}])

fiscal_pillar = pd.DataFrame([{
    'short_name': 'fiscal_header',
    'long_name': 'FISCAL PILLAR (17%)',
    'coefficient': '',
    'Z-score Value': '',
    'Rating (notches)': ''
}])

external_pillar = pd.DataFrame([{
    'short_name': 'ext_header',
    'long_name': 'EXTERNAL PILLAR (14%)',
    'coefficient': '',
    'Z-score Value': '',
    'Rating (notches)': ''
}])

final_pillar = pd.DataFrame([{
    'short_name': 'final_header',
    'long_name': 'SOVEREIGN CREDIT RATING',
    'coefficient': '',
    'Z-score Value': '',
    'Rating (notches)': ''
}])

## Slice DataFrame based on where the insertions should go
df1 = short_table_df.iloc[:1]   # includes const row
df2 = short_table_df.iloc[1:4]  # economy factors. includes rows 1-3, excludes 4
df3 = short_table_df.iloc[4:7]  # institutional factors, includes rows 4-6, excludes 7
df4 = short_table_df.iloc[7:9]   # fiscal factors, includes rows 7-9, excludes 9
df5 = short_table_df.iloc[9:12]   # external factors, includes rows 9-12, excludes 12
df6 = short_table_df.iloc[12:14] #Sovereign Credit Rating, includes rows 12-13, excludes 14

## Combine everything in order to make short_table_df once again

short_table_df = pd.concat(
    [df1, economy_pillar, df2, institutions_pillar, df3, fiscal_pillar, df4, external_pillar, df5, final_pillar,df6 ],
    ignore_index=True
)

# Change column names to make more presentable

short_table_df = short_table_df.rename(columns={'long_name': 'Factor'})

# Inserting override logic to allow user interaction. HARDEST PART!!

## Load Overrides Based on Country-Year

## To do this we need to set up to connect to google sheets

@st.cache_resource
def init_gsheets_client():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds_info = dict(st.secrets["gcp_service_account"])
    creds_info["private_key"] = creds_info["private_key"].replace("\\n", "\n")
    creds = Credentials.from_service_account_info(creds_info, scopes=scope)
    return gspread.authorize(creds)

client = init_gsheets_client()

## (the below segment is the old code along with explainers...)

## scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
#Defines the authorization scopes — i.e., what permissions your app is requesting from Google. 
#Allows access to read/write Google Sheets data
#Allows access to open the sheet (via Google Drive), even if it's not explicitly listed in your Drive UI
#These are required for gspread to function correctly.

## creds_info = dict(st.secrets["gcp_service_account"]) #--> config object st.secrets. need to convert to a DICT

## creds_info["private_key"] = creds_info["private_key"].replace("\\n", "\n") # Repair the escaped-newline issue

#Pulls your service account credentials from Streamlit’s secrets.toml file, where you’ve stored the [gcp_service_account] block.

#creds_dict is now a Python dictionary containing your private key, client email, etc.

##creds = Credentials.from_service_account_info(creds_info, scopes = scope) #important to put scopes = " ". otherwise position wrong

#Converts the credentials dictionary into a usable OAuth2 credentials object that gspread can use to authenticate
#Think of it as logging in with the service account and telling Google what scopes (permissions) your app wants.

## client = gspread.authorize(creds)
#Uses the credentials to create a gspread client — this is your authenticated connection to Google Sheets
#You’ll use client to open any sheet, read, write, or update data.

## sheet_short = client.open("analyst_overrides_short")
#Opens the Google Sheet named "analyst_overrides_short"
#Note that streamlit "sees" this as the entire spreadsheet since we didn't specify specific tab
#Now sheet_short is a live object that lets you read from or write to that google sheet

#With the connection established. Let us load the analyst overrides into our short_table_df
## override_df = load_override_from_gsheet(sheet_short, selected_name, selected_year)
#what this function does is looks at the google sheet object (sheet_short in this case)
#uses selected_name to find the relevant country tab (cos i named each tab with a different country name)
#within the country tab, searches for overrides in a specific year (selected_year) "short_name", "Adjustment", "Analyst Comment"
#calls this out as a df called override_df

sheet_short = client.open("analyst_overrides_short")

@st.cache_data
def fetch_overrides(country: str, year: int):
    sheet_short = client.open("analyst_overrides_short")
    return load_override_from_gsheet(sheet_short, country, year)

override_df = fetch_overrides(selected_name, selected_year)

## Loading block complete ##

## Merge overrides into the main df
short_table_df = pd.merge(short_table_df, override_df, on="short_name", how="left")
short_table_df["Adjustment"] = pd.to_numeric(short_table_df["Adjustment"], errors="coerce").fillna(0)
short_table_df["Analyst Comment"] = short_table_df["Analyst Comment"].fillna("")

## Detour to add model predicted rating (letter) in the model_rating (row) and Adjustment (col) spot

### Step 1: Convert rating_index into a dictionary for quick mapping to letter rating

rating_dict = dict(zip(rating_index['Numeric'], rating_index['Credit Rating'])) #zip pairs the two columns row by row to help make into a dict

### Step 2: Map numeric model rating to corresponding letter rating in rating_dict
clamped_rating = min(22,max(1,round(model_rating))) # to be force model numeric rating to be between 1-22 rating scale
letter_rating = rating_dict.get(clamped_rating,'N/A')

### Step 3: place letter_rating in the model_rating (row) and Analyst comment (col) spot
short_table_df.loc[short_table_df["short_name"] == "predicted_rating", "Analyst Comment"] = letter_rating

## Sum up all adjustments except for the predicted_rating row
adj_sum = short_table_df.loc[
    (short_table_df["short_name"] != "predicted_rating") & (short_table_df["short_name"] != ""),
    "Adjustment"].sum()
#first part creates a boolean of rows where short name is not equals predicted rating
#and excludes headers and blank rows
#it then sums up the values found in the adjustment column
#saves the sum of analyst adjustments as variable adj_sum

## Calculate adjusted rating
adjusted_rating = model_rating + adj_sum #recall model_rating was calculated at the outset when making short_table_df

## Update adj_sum into the model_rating row and adjustment column
short_table_df.loc[short_table_df["short_name"] == "predicted_rating", "Adjustment"] = adj_sum

## Create final_rating row

final_row = pd.DataFrame([{
    'short_name': 'final_rating',
    'Factor': 'LS Final Rating',
    'coefficient': '',
    'Z-score Value': '',
    'Rating (notches)': adjusted_rating
}])

## Append it to the bottom
short_table_df = pd.concat([short_table_df, final_row], ignore_index=True)

## Update final_rating row and analyst comment column with adjusting_rating letter
clamped_rating_adj = min(22,max(1,round(adjusted_rating))) # to be force model numeric rating to be between 1-22 rating scale
letter_rating_adj = rating_dict.get(clamped_rating_adj,'N/A')
short_table_df.loc[short_table_df["short_name"] == "final_rating", "Analyst Comment"] = letter_rating_adj

## USe ST metric to show adjusted rating and public credit ratings right at the top

### First clean up the public_rating_index file. This needs to be hardcoded in excel due to how bbg works!

for col in ['moodys', 's&p', 'fitch']:
    # 1) convert NaNs → 'NR'
    public_rating_index[col] = public_rating_index[col].fillna('NR')
    # 2) drop any trailing 'u' (e.g. 'Baa1u' → 'Baa1')
    public_rating_index[col] = public_rating_index[col].str.replace(r'u$', '', regex=True)
    #r'u$' is a regular expression. r is string. u is the literal "u".$ means anchored to end of string.
    #regex = True tells Pandas to treat the first arg as a regulat expression rather than a string

### And then we pull out select_country and the relevant rating agency (header name) to get the public ratings
### we then display them using the st.metric function from streamlit

col1, col2, col3, col4 = st.columns(4)
col1.metric(label = "LS",value = letter_rating_adj)
col2.metric(label = "S&P",value = public_rating_index.loc[public_rating_index["name"] == selected_name,"s&p"].values[0])
col3.metric(label = "Moody's",value = public_rating_index.loc[public_rating_index["name"] == selected_name,"moodys"].values[0])
col4.metric(label = "Fitch",value = public_rating_index.loc[public_rating_index["name"] == selected_name,"fitch"].values[0])

# subheader to appear after drop down and before rating table
st.subheader("Main Credit Rating Table (11 Factor Model)")

# Initialize AgGrid to create interactive table in 

## Basic column options
gb = GridOptionsBuilder.from_dataframe(short_table_df)

## Apply to all columns
gb.configure_default_column(
    editable=False, #cannot edit
    sortable=False, #cannot sort
    filter=False, #cannot filter
    resizable=False, #cannot resize
    suppressHeaderMenuButton=True #hide the 3 dots button for me. FINALLY!
)

## Using JsCode to effect how different cell numerical values are perceived
## Note that underlying data in the df is not changed

### Two Requests
### 1. “If the value is missing, show blank. If it’s a letter like 'A', just display it as-is — don't try to force into a numeric”
### 2. Show numeric value as 1 decimal

combined_formatter = JsCode("""
function(params) {
  const v = params.value;
  // 1) Blank out null/undefined/empty
  if (v === undefined || v === null || v === '') {
    return '';
  }
  // 2) If numeric, show one decimal
  if (!isNaN(v)) {
    return parseFloat(v).toFixed(2);
  }
  // 3) Otherwise (e.g. letter ratings), just display as string
  return v.toString();
}
""")

hide_na_formatter = JsCode("""
function(params) {
    return params.value === undefined || params.value === null ? '' : params.value.toString();
}
""") ### “If the value is missing, show blank. If it’s a letter like 'A', just display it as-is — don't try to force into a numeric”

hide_zero_formatter = JsCode("""
function(params) {
  const v = params.value;
  // 1) blank out null/undefined and zero
  if (v === undefined || v === null || v === 0) {
    return '';
  }
  // 2) if it’s numeric, show two decimals
  if (!isNaN(v)) {
    return parseFloat(v).toFixed(1);
  }
  // 3) otherwise (e.g. letter), just render as string
  return v.toString();
}
""")

### This handles:
### null / undefined → blank
### numbers → rounded to 1 decimal
### everything else → cast to string

## Apply column by column configuration in df...
gb.configure_column("short_name", hide=True)
gb.configure_column("Factor", valueFormatter=combined_formatter, cellStyle=JsCode("""
      function(params) {
        return { 'font-weight': 'bold' };
      }
    """)) ## the cell style here makes all the fonts bold..
gb.configure_column("coefficient", valueFormatter=combined_formatter)
gb.configure_column("Z-score Value", valueFormatter=combined_formatter)
gb.configure_column("Rating (notches)", valueFormatter=combined_formatter, cellStyle=JsCode("""
      function(params) {
        const id = params.data.short_name;
        if (id === 'predicted_rating' || id === 'final_rating') {
          return {
            color: 'black',
            'font-weight': 'bold'
          };
        }
        return null;  // leave all other rows with their default style
      }
    """))

editable_callback = JsCode("""
  function(params) {
    const id = params.data.short_name;
    // editable only if:
    //  • it's not a header (id !== '')
    //  • it's not the model output rows
    return id !== '' && id !== 'predicted_rating' && id !== 'final_rating';
  }
""") ## creating this special object to put into the editable argument for the last 2 col. as dont want people editting header or final score

gb.configure_column("Adjustment", valueFormatter=hide_zero_formatter,editable=editable_callback, filter=False, headerClass="ag-header-cell-label-left",
                    cellClass="ag-left-aligned-cell",cellStyle=JsCode("""
  function(params) {
    const id = params.data.short_name;
    if (id === 'predicted_rating' || id === 'final_rating') {
      return { color: 'black', 'font-weight': 'bold' };
    }
    return { color: 'blue',  'font-weight': 'normal' };
  }
"""))
gb.configure_column("Analyst Comment", valueFormatter=hide_na_formatter,editable=editable_callback,maxWidth=488,minWidth=488,
                    cellStyle=JsCode("""
  function(params) {
    const id = params.data.short_name;
    if (id === 'predicted_rating' || id === 'final_rating') {
      return { color: 'black', 'font-weight': 'bold' };
    }
    return { color: 'blue',  'font-weight': 'normal' };
  }
"""))
### cellClass="ag-right-aligned-cell" aligns the column contents either left or right
### headerClass="ag-header-cell-label-left" aligns header text to the left or right
### cell style uses java script to affect how text looks in the column.

## Attempt to style the table to make it look Loomis Excel Like

### Define your overrides in a dict and pass to AgGrid
### this overrides the theme

custom_css_override = {
    # header cell label (the text container)
    ".ag-header-cell-label": {
        "background-color": "#1A3B73 !important",
        "color":            "white !important",
        "font-weight":      "bold !important",
        #"font-size":        "16px !important" #messes wtih alignment..
    },
    # the very header row wrapper (fills behind the labels)
    ".ag-header": {
        "background-color": "#1A3B73 !important",
    },
}

# 1) After all your gb.configure_… calls:

LS_gridOptions = gb.build()

### LS_gridOptions["domLayout"] = "print" 
### If you don’t need the internal grid scrollbars at all, you can tell AG-Grid to render all columns in one shot and size to fit with this..

# 2) Attach your multi‐row styling! add getRowStyle to bold and highlight the "predicted_rating" row
LS_gridOptions["getRowStyle"] = JsCode("""
  function(params) {
    // list of the exact Factor values you want to style
    const headers = [
      "REAL ECONOMY PILLAR (25%)",
      "MONETARY & INSTITUTIONS PILLAR (44%)",
      "FISCAL PILLAR (17%)",
      "EXTERNAL PILLAR (14%)",
      "SOVEREIGN CREDIT RATING"
    ];
    // if this row’s Factor is one of the headers, return a style object
    if (headers.includes(params.data.Factor)) {
      return {
        "font-weight":      "bold",
        "background-color": "#B6CEE4"   // light tint—change as you like
      };
    }
    return null;  // otherwise use default styling
  }
""")

## Render the table using AgGrid (and pass in your overrides injecting the CSS injection)

grid_response = AgGrid(
    short_table_df,
    gridOptions=LS_gridOptions, # <- use the variable from the row options gb build
    custom_css = custom_css_override,
    allow_unsafe_jscode=True,
    update_mode='VALUE_CHANGED',  #necessary to capture edits
    fit_columns_on_grid_load=False,# we’re sizing to contents instead
    columns_auto_size_mode=ColumnsAutoSizeMode.FIT_CONTENTS, #columns size to fit contents
    suppressColumnVirtualisation=True,    # measure off-screen columns too
    theme='alpine',
    height=500,  # manually control table height without scrolling
    
)

### What is grid_response that i created using AgGrid()? 
### It’s a dict (or more precisely, a Box object — behaves like a dict) with multiple keys that give you access to:
### 'data' -->	pd.DataFrame --> The updated DataFrame after user edits
### 'selected_rows' --> list[dict] --> 	A list of rows the user selected (if row selection is enabled)
### 'column_state' --> list[dict] --> The state of columns (e.g. width, sort order)
### 'rowData' --> list[dict] --> Raw row data as a list of dictionaries (alternative to data)

## Captures edits made by user in grid
updated_df = grid_response["data"] #extracts the updated DataFrame after user edits from AgGrid (adjustment and comments col)
updated_df["Adjustment"] = pd.to_numeric(updated_df["Adjustment"], errors="coerce").fillna(0) #safety layer to ensure only numeric captured
#errors = coerce means you dont crash the app if non numeric. just input nan value. which we then turn to zero!

# Create formatted excel file for export
export_short_df = updated_df.drop(columns=['short_name'])

def generate_custom_export(
    df: pd.DataFrame,
) -> BytesIO:
    wb = Workbook()
    ws = wb.active

    # 1) Insert 5 blank rows at the top
    ws.insert_rows(idx=1, amount=5)

    # 2) Populate A1/A2/A3/A4
    ws["A1"] = "Country"
    ws["A2"] = selected_name
    ws["A3"] = "Year"
    ws["A4"] = selected_year

    # 3) Write DataFrame header at row 6, data from row 7 onward
    header_row = 6
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=col_idx, value=col_name)
    for r, row in enumerate(df.itertuples(index=False), start=header_row+1):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    # 4) Define styles
    dark_blue   = PatternFill("solid", fgColor="FF1A3B73")
    gray_fill   = PatternFill("solid", fgColor="FFF2F2F2")
    light_blue  = PatternFill("solid", fgColor="FFB6CEE4")
    gray_e9     = PatternFill("solid", fgColor="FFE9E9EB")
    tint_da     = PatternFill("solid", fgColor="FFDAEEF3")
    white_font  = Font(color="FFFFFFFF")
    bold_font   = Font(bold=True)
    header_font = Font(bold=True, color="FFFFFFFF")  # bold + white-text
    blue_font   = Font(color="FF0000FF")
    left_align  = Alignment(horizontal="left")
    thin_side   = Side(style="thin")
    bold_white_font = Font(bold=True, color="FFFFFFFF")

    def make_border(top=False, bottom=False, left=False, right=False):
        return Border(
            top    = thin_side if top    else Side(style=None),
            bottom = thin_side if bottom else Side(style=None),
            left   = thin_side if left   else Side(style=None),
            right  = thin_side if right  else Side(style=None),
        )

    ncols      = df.shape[1]
    data_start = header_row + 1
    data_end   = data_start + len(df) - 1

    # 5) Style A2/A4: gray fill, bold font
    for coord in ("A2", "A4"):
        ws[coord].fill      = gray_fill
        ws[coord].font      = bold_font
        ws[coord].alignment = left_align

    # 6) Bold entire Column A and auto-fit its width
    max_w = 0
    for r in range(1, data_end + 1):
        c = ws.cell(row=r, column=1)
        c.font = bold_font
        length = len(str(c.value or ""))
        if length > max_w:
            max_w = length
    ws.column_dimensions["A"].width = max_w + 2

    # 7) Style A1/A3: dark blue fill, white font
    for coord in ("A1", "A3"):
        ws[coord].fill      = dark_blue
        ws[coord].font      = bold_white_font # ensure white font and bolded
        ws[coord].alignment = left_align

    # 8) Header row (A4:F4): dark blue fill, bold+white text, borders
    for col in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill   = dark_blue
        cell.font   = header_font
        cell.border = make_border(
            top   = True,
            left  = (col == 1),
            right = (col == ncols)
        )

    # 9) Number formats & zero suppression
    for r in range(data_start, data_end + 1):
        for col_idx in (2, 3, 4):
            c = ws.cell(row=r, column=col_idx)
            if isinstance(c.value, (int, float)):
                c.number_format = "0.00"
        e = ws.cell(row=r, column=5)
        if isinstance(e.value, (int, float)):
            if e.value == 0:
                e.value = None
            else:
                e.number_format = "0.0"

    # 10A) Blue font in E7:E22 (or up to data_end)
    for r in range(data_start, min(data_end, 22) + 1):
        ws.cell(row=r, column=5).font = blue_font
    
    # 10B) Blue font in F7:F22 (or up to data_end)
    for r in range(data_start, min(data_end, 22) + 1):
        ws.cell(row=r, column=6).font = blue_font

    # 11) Shade specific rows A8:F8, A12:F12, A16:F16, A19:F19, A23:F23
    for r in (8, 12, 16, 19, 23):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).fill = light_blue

    # 12) Fill A7 with light gray
    ws["A7"].fill = gray_e9

    # 13) Fill A9–A11, A13–A15, A17–A18, A20–A22 with tint_da
    for start, end in ((9, 11), (13, 15), (17, 18), (20, 22)):
        for r in range(start, end + 1):
            ws.cell(row=r, column=1).fill = tint_da

    # 14) Bold fonts in D24, D25, E24, F24, F25
    for (r, c) in ((24, 4), (25, 4), (24, 5), (24, 6), (25, 6)):
        ws.cell(row=r, column=c).font = bold_font

    # 15) Fill A24:F25 with gray_fill
    for r in (24, 25):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).fill = gray_fill

    # 16) Fixed widths for columns A–F
    # Define desired widths in Excel’s column-width units
    fixed_widths = {
    1: 25.0,  # A  
    2: 10,   # B
    3: 12,  # C
    4: 14.5,  # D
    5: 11.5,     # E
    6: 57      # F
    }
    
    for col_idx, width in fixed_widths.items():
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = width
    
    # 17) Full outside border from A1 to F25
    for r in range(1, 26):
        for c in range(1, ncols + 1):
            if r in (1, 25) or c in (1, ncols):
                sides = {
                    "top":    (r == 1),
                    "bottom": (r == 25),
                    "left":   (c == 1),
                    "right":  (c == ncols)
                }
                ws.cell(row=r, column=c).border = make_border(**sides)

    # 18) Save to BytesIO
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

excel_data = generate_custom_export(export_short_df)

# Put the Save + Export buttons side by side
# carve the page into 3 chunks: 
#  • 1 unit for btn1 
#  • 1 unit for btn2 
#  • 6 units of blank space

save_col_short, export_col_short, blank_col_short = st.columns([2, 2, 6])

with save_col_short:
    if st.button("💾 Save Analyst Overrides",key="short_save"):
        # Save only the override columns (factor-level edits) to a file
        columns_to_save = ["short_name", "Adjustment", "Analyst Comment"]
        updated_subset = updated_df[columns_to_save]
        
        # Use the full Google Sheet, then pass selected_name to target the right tab
        save_override_to_gsheet(sheet_short, updated_subset, selected_name, selected_year)

        # clear only the cache for fetch_overrides
        fetch_overrides.clear()
        
        st.success("✅ Overrides saved and rating updated.")
        st.rerun() #rerun entire script from top to bottom so analyst can see update immediately

with export_col_short:
    st.download_button(
    label="📥 Export to Excel (Formatted)",
    key = "short_excel",
    data=excel_data,
    file_name="main_rating_table.xlsx",
    mime="application/vnd.openxmlformats-officedocument-spreadsheetml.sheet"
)

##### HERE WE START ON THE LONG TABLE -------------------------------------####

# subheader to appear before dropdown
st.subheader("Supplementary Credit Rating Table (Constituent Variables)")

# Select Row based on year and country
selected_row_raw = df_raw[(df_raw['name'] == selected_name) & (df_raw['year'] == selected_year)]
selected_row_transform = df_transform[(df_transform['name'] == selected_name) & (df_transform['year'] == selected_year)]

# Use variable_index (variable name file from excel) to form LHS of long_table_df

variable_index_long = variable_index.copy()
long_table_df = variable_index_long

# Transpose the selected country / year row from both the raw and transform data frames

long_table_df_raw = selected_row_raw.T.reset_index() #T is transpose. Reset Index just prevents variable name from being index
long_table_df_raw.columns = ['short_name', 'Raw Value'] #rename columns

long_table_df_transform = selected_row_transform.T.reset_index() #T is transpose. Reset Index just prevents variable name from being index
long_table_df_transform.columns = ['short_name', 'Z-score Value'] #rename columns

# Have to tweak the row names in long_table_raw first col so that single variable factors get renamed as factor for LHS join

long_table_df_raw['short_name'] = long_table_df_raw['short_name'].replace({
    'ngdp_pc': 'wealth_factor',
    'ngdp': 'size_factor',
    'growth_avg': 'growth_factor',
    'inf_avg': 'inflation_factor',
    'gov_debt_gdp': 'govdebt_factor',
    'cab_avg': 'extperf_factor',
    'reserve_fx': 'reservestatus_factor'
})


# merge all 3 table into a single long_table_df using left join

# merge long_table_df with long_table_df_raw
long_table_df = pd.merge(long_table_df, long_table_df_raw, on = 'short_name', how='left')

# then merge the combined table with long_table_df_transform to get the final table
long_table_df = pd.merge(long_table_df, long_table_df_transform, on = 'short_name', how='left')

# Manually Insert Section Headers

## Build section header rows
economy_pillar = pd.DataFrame([{
    'short_name': 'eco_header',
    'long_name': 'REAL ECONOMY PILLAR (25%)',
    'description': '',
    'Raw Value': '',
    'Z-score Value': ''
}])

institutions_pillar = pd.DataFrame([{
    'short_name': 'insti_header',
    'long_name': 'MONETARY & INSTITUTIONS PILLAR (44%)',
    'description': '',
    'Raw Value': '',
    'Z-score Value': ''
}])

fiscal_pillar = pd.DataFrame([{
    'short_name': 'fiscal_header',
    'long_name': 'FISCAL PILLAR (17%)',
    'description': '',
    'Raw Value': '',
    'Z-score Value': ''
}])

external_pillar = pd.DataFrame([{
    'short_name': 'ext_header',
    'long_name': 'EXTERNAL PILLAR (14%)',
    'description': '',
    'Raw Value': '',
    'Z-score Value': ''
}])

## Slice DataFrame based on where the insertions should go
df1 = long_table_df.iloc[:1]   # includes const row
df2 = long_table_df.iloc[1:4]  # economy factors. includes rows 1-3, excludes 4
df3 = long_table_df.iloc[4:15]  # institutional factors, includes rows 4-14, excludes 15
df4 = long_table_df.iloc[15:20]   # fiscal factors, includes rows 15-19, excludes 20
df5 = long_table_df.iloc[20:25]   # external factors, includes rows 20-24, excludes 25

## Combine everything in order to make long_table_df once again

long_table_df = pd.concat(
    [df1, economy_pillar, df2, institutions_pillar, df3, fiscal_pillar, df4, external_pillar, df5],
    ignore_index=True
)

# Change column names to make more presentable

long_table_df = long_table_df.rename(columns={'long_name':'Factor','description': 'Constituent Variables'})

# Inserting override logic to allow user interaction. HARDEST PART!!

#already ran the authorization block of code to google sheets above. now we just use client object to open a new sheet
sheet_long = client.open("analyst_overrides_long")

## With the connection established. Let us load the analyst overrides into our long_table_df

# Do similar caching to short_table above. only load unique country year combination once unless there is edit.
@st.cache_data
def fetch_overrides_long(country: str, year: int):
    sheet_long = client.open("analyst_overrides_long")
    return load_override_from_gsheet(sheet_long, country, year)

override_df_long = fetch_overrides_long(selected_name, selected_year)

## override_df_long = load_override_from_gsheet(sheet_long, selected_name, selected_year)
#what this function does is looks at the google sheet object (sheet_long in this case)
#uses selected_name to find the relevant country tab (cos i named each tab with a different country name)
#within the country tab, searches for overrides in a specific year (selected_year) "short_name", "Adjustment", "Analyst Comment"
#calls this out as a df called override_df_long

## Merge overrides into the main df
long_table_df = pd.merge(long_table_df, override_df_long, on="short_name", how="left")
long_table_df["Adjustment"] = pd.to_numeric(long_table_df["Adjustment"], errors="coerce").fillna(0)
long_table_df["Analyst Comment"] = long_table_df["Analyst Comment"].fillna("")

## create sums for the various constituent variables that roll up to the factor level

### 1)sum for default sub-factors
default_subfactors = ["default_hist", "default_decay"]

#### Compute subtotal
default_sum = long_table_df.loc[long_table_df["short_name"].isin(default_subfactors), "Adjustment"].sum()

#### Assign it to the default_factor row
long_table_df.loc[long_table_df["short_name"] == "default_factor", "Adjustment"] = default_sum

### 2)sum for governance sub-factors
governance_subfactors = ["voice_acct", "pol_stab", "gov_eff","reg_qual","rule_law","cont_corrupt"]

#### Compute subtotal
governance_sum = long_table_df.loc[long_table_df["short_name"].isin(governance_subfactors), "Adjustment"].sum()

#### Assign it to the governance_factor row
long_table_df.loc[long_table_df["short_name"] == "governance_factor", "Adjustment"] = governance_sum

### 3)sum for fiscal performance sub-factors
fiscal_subfactors = ["fb_avg", "gov_rev_gdp", "ir_rev"]

#### Compute subtotal
fiscal_sum = long_table_df.loc[long_table_df["short_name"].isin(fiscal_subfactors), "Adjustment"].sum()

#### Assign it to the fiscal_factor row
long_table_df.loc[long_table_df["short_name"] == "fiscalperf_factor", "Adjustment"] = fiscal_sum

### 4)sum for FX reserves factor
fx_subfactors = ["reserve_gdp", "import_cover"]

#### Compute subtotal
fx_sum = long_table_df.loc[long_table_df["short_name"].isin(fx_subfactors), "Adjustment"].sum()

#### Assign it to the reservebuffer_factor row
long_table_df.loc[long_table_df["short_name"] == "reservebuffer_factor", "Adjustment"] = fx_sum

# Initialize AgGrid to create interactive table in 

## Basic column options
gb_long = GridOptionsBuilder.from_dataframe(long_table_df)

## Apply to all columns
gb_long.configure_default_column(
    editable=False, #cannot edit
    sortable=False, #cannot sort
    filter=False, #cannot filter
    resizable=False, #cannot resize
    suppressHeaderMenuButton=True #hide the 3 dots button for me. FINALLY!
)

## Define formatters which basically controls how values display in the column
## Define styles which basically shows how values look in a column (bold etc etc)
## editable call_back (restrictions) that i can pass into the editable arg to control which cells can or cannot be edited

combined_formatter_long = JsCode("""
function(params) {
  const v = params.value;
  // 1) Blank out null/undefined/empty
  if (v === undefined || v === null || v === '') {
    return '';
  }
  // 2) If numeric, show one decimal
  if (!isNaN(v)) {
    return parseFloat(v).toFixed(2);
  }
  // 3) Otherwise (e.g. letter ratings), just display as string
  return v.toString();
}
""")

rawvalue_formatter_long = JsCode("""
function(params) {
  const v  = params.value;
  const id = params.data.short_name;

  // 1) Handle null/undefined/empty
  if (v === undefined || v === null || v === '') {
    // show “–” for these specific headers
    const dashRows = [
      'default_factor',
      'governance_factor',
      'fiscalperf_factor',
      'reservebuffer_factor'
    ];
    return dashRows.includes(id) ? '–' : '';
  }

  // 2) Parse number
  const num = parseFloat(v);
  if (isNaN(num)) {
    // if it wasn’t numeric, just show it as text
    return v.toString();
  }

  // 3) Formatting by row key
  switch (id) {
    case 'wealth_factor':
      // comma thousands, no decimals
      return num.toLocaleString(undefined, {
        minimumFractionDigits: 0,
        maximumFractionDigits: 0
      });

    case 'default_hist':
    case 'reservestatus_factor':
      // integer (0 or 1)
      return num.toFixed(0);

    default:
      // everything else: one decimal place
      return num.toFixed(1);
  }
}
""")

hide_na_formatter_long = JsCode("""
function(params) {
    return params.value === undefined || params.value === null ? '' : params.value.toString();
}
""") ### “If the value is missing, show blank. If it’s a letter like 'A', just display it as-is — don't try to force into a numeric”

hide_zero_formatter_long = JsCode("""
function(params) {
  const v = params.value;
  // 1) blank out null/undefined and zero
  if (v === undefined || v === null || v === 0) {
    return '';
  }
  // 2) if it’s numeric, show two decimals
  if (!isNaN(v)) {
    return parseFloat(v).toFixed(1);
  }
  // 3) otherwise (e.g. letter), just render as string
  return v.toString();
}
""")

purple_values_style = JsCode("""
function(params) {
  const id = params.data.short_name;
  const purpleIds = [
    'default_hist',
    'default_decay',
    'voice_acct',
    'pol_stab',
    'gov_eff',
    'reg_qual',
    'rule_law',
    'cont_corrupt',
    'fb_avg',
    'gov_rev_gdp',
    'ir_rev',
    'reserve_gdp',
    'import_cover'
  ];
  if (purpleIds.includes(id)) {
    return { color: '#B21740' };
  }
  return null;
}
""")

purple_description_style = JsCode("""
function(params) {
  const id = params.data.short_name;
  const purpleIds = [
    'default_hist',
    'default_decay',
    'voice_acct',
    'pol_stab',
    'gov_eff',
    'reg_qual',
    'rule_law',
    'cont_corrupt',
    'fb_avg',
    'gov_rev_gdp',
    'ir_rev',
    'reserve_gdp',
    'import_cover'
  ];
  if (purpleIds.includes(id)) {
    // purple, normal weight
    return { color: '#B21740', 'font-weight': 'normal' };
  }
  // all other rows: black, bold
  return { color: 'black', 'font-weight': 'bold' };
}
""")

adjustment_style = JsCode("""
function(params) {
  const id = params.data.short_name;
  // list of rows to render in blue + bold
  const blueIds = [
    'wealth_factor',
    'size_factor',
    'growth_factor',
    'inflation_factor',
    'default_factor',
    'governance_factor',
    'fiscalperf_factor',
    'govdebt_factor',
    'extperf_factor',
    'reservebuffer_factor',
    'reservestatus_factor'
  ];
  if (blueIds.includes(id)) {
    return {
      color: '#0000FF',
      'font-weight': 'bold'
    };
  }
  // everything else: maroon-ish, normal weight
  return {
    color: '#B21740',
    'font-weight': 'normal'
  };
}
""")

analyst_style = JsCode("""
function(params) {
  const id = params.data.short_name;
  const blueIds = [
    'wealth_factor',
    'size_factor',
    'growth_factor',
    'inflation_factor',
    'default_factor',
    'governance_factor',
    'fiscalperf_factor',
    'govdebt_factor',
    'extperf_factor',
    'reservebuffer_factor',
    'reservestatus_factor'
  ];
  if (blueIds.includes(id)) {
    return { color: '#0000FF' };      // blue, no bold
  }
  return { color: '#B21740' };        // maroon-ish, no bold
}
""")

factor_style_long = JsCode("""
function(params) {
  const style = {};
  // header sentinels to skip
  const headers = [
    "eco_header",
    "insti_header",
    "fiscal_header",
    "ext_header",
    "final_header"
  ];
  // Shade every non-header row
  if (!headers.includes(params.data.short_name)) {
    style['background-color'] = '#DAEEF3';
  }
  // Bold all text
  style['font-weight'] = 'bold';
  return style;
}
""")

editable_criteria_adjustment = JsCode("""
function(params) {
  const id = params.data.short_name;
  // list of short_name values that should NOT be editable
  const locked = [
    '',                  // blank header rows
    'const',
    'default_factor',
    'governance_factor',
    'fiscalperf_factor',
    'reservebuffer_factor'
  ];
  // return false (lock) when id is in our locked list; true otherwise
  return !locked.includes(id);
}
""")

editable_criteria_analyst = JsCode("""
function(params) {
  const id = params.data.short_name;
  // lock only blank header rows and the 'const' row
  if (id === '' || id === 'const') {
    return false;
  }
  return true;
}
""")

## Apply column by column configuration in df...

gb_long.configure_column("short_name", hide=True)
gb_long.configure_column("Factor", valueFormatter=combined_formatter_long, cellStyle=factor_style_long)
gb_long.configure_column("Constituent Variables", valueFormatter=combined_formatter_long, cellStyle=purple_description_style)
gb_long.configure_column("Raw Value", valueFormatter=rawvalue_formatter_long, cellStyle = purple_values_style)
gb_long.configure_column("Z-score Value", valueFormatter=combined_formatter_long, cellStyle = purple_values_style)

## Make adjustment and analyst rationale columns editable

gb_long.configure_column("Adjustment",valueFormatter=hide_zero_formatter_long, cellStyle = adjustment_style,
                         editable = editable_criteria_adjustment, filter=False, headerClass="ag-header-cell-label-left",
                         cellClass="ag-left-aligned-cell")
gb_long.configure_column("Analyst Comment",valueFormatter=hide_na_formatter_long, cellStyle = analyst_style, 
                         editable = editable_criteria_analyst,minWidth=488)

## Now we apply the color schemes to the grid

custom_css_override_long = {
    # header cell label (the text container)
    ".ag-header-cell-label": {
        "background-color": "#1A3B73 !important",
        "color":            "white !important",
        "font-weight":      "bold !important",
        #"font-size":        "16px !important"
    },
    # the very header row wrapper (fills behind the labels)
    ".ag-header": {
        "background-color": "#1A3B73 !important"
    },
}

LS_gridOptions_long = gb_long.build()
LS_gridOptions_long["getRowStyle"] = JsCode("""
  function(params) {
    // list of the exact Factor values you want to style
    const headers = [
      "REAL ECONOMY PILLAR (25%)",
      "MONETARY & INSTITUTIONS PILLAR (44%)",
      "FISCAL PILLAR (17%)",
      "EXTERNAL PILLAR (14%)",
      "SOVEREIGN CREDIT RATING"
    ];
    // if this row’s Factor is one of the headers, return a style object
    if (headers.includes(params.data.Factor)) {
      return {
        "font-weight":      "bold",
        "background-color": "#B6CEE4"   // light tint—change as you like
      };
    }
    return null;  // otherwise use default styling
  }
""")


## Finally we initialize the grid

grid_response_long = AgGrid(
    long_table_df,
    gridOptions=LS_gridOptions_long,
    custom_css = custom_css_override_long,
    allow_unsafe_jscode=True,
    update_mode='VALUE_CHANGED',  #necessary to capture edits
    fit_columns_on_grid_load=False,# we’re sizing to contents instead
    columns_auto_size_mode=ColumnsAutoSizeMode.FIT_CONTENTS, #columns size to fit contents
    suppressColumnVirtualisation=True,    # measure off-screen columns too
    theme='alpine',
    height=500,  # manually control table height without scrolling
    )

## Captures edits made by user in grid
updated_df_long = grid_response_long["data"] #extracts the updated DataFrame after user edits from AgGrid (adjustment and comments col)
updated_df_long["Adjustment"] = pd.to_numeric(updated_df_long["Adjustment"], errors="coerce").fillna(0) #safety layer to ensure only numeric captured
#errors = coerce means you dont crash the app if non numeric. just input nan value. which we then turn to zero!

# Create formatted excel file for export
export_long_df = updated_df_long.drop(columns=['short_name'])

def generate_custom_export_long(
    df: pd.DataFrame,
) -> BytesIO:
    wb = Workbook()
    ws = wb.active

    # 1) Insert 5 blank rows
    ws.insert_rows(idx=1, amount=5)

    # 2) Populate A1–A4
    ws["A1"] = "Country"
    ws["A2"] = selected_name
    ws["A3"] = "Year"
    ws["A4"] = selected_year

    # 3) Write headers at row 6, data from row 7
    header_row = 6
    for ci, col in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=ci, value=col)
    data_start = header_row + 1
    for ri, row in enumerate(df.itertuples(index=False), start=data_start):
        for ci, v in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value=v)
    data_end = data_start + len(df) - 1

    # 4) Define styles
    dark_blue  = PatternFill("solid", fgColor="FF1A3B73")
    gray_fill  = PatternFill("solid", fgColor="FFF2F2F2")
    light_blue = PatternFill("solid", fgColor="FFB6CEE4")
    gray_e9    = PatternFill("solid", fgColor="FFE9E9EB")
    tint_da    = PatternFill("solid", fgColor="FFDAEEF3")
    maroon     = Font(color="FFB21740")
    white      = Font(color="FFFFFFFF")
    bold       = Font(bold=True)
    bold_white = Font(bold=True, color="FFFFFFFF")
    blue_font  = Font(color="FF0000FF")
    left       = Alignment(horizontal="left")
    right      = Alignment(horizontal="right")
    thin       = Side(style="thin")

    def mk_border(top=False, bottom=False, left_b=False, right_b=False):
        return Border(
            top    = thin if top    else Side(style=None),
            bottom = thin if bottom else Side(style=None),
            left   = thin if left_b else Side(style=None),
            right  = thin if right_b else Side(style=None),
        )

    # 5) Bold all col A
    for r in range(1, data_end + 1):
        ws.cell(row=r, column=1).font = bold

    # 6) Shade A1/A3 dark blue + white font
    for coord in ("A1","A3"):
        ws[coord].fill = dark_blue
        ws[coord].font = bold_white

    # 7) Shade A2/A4 gray + bold
    for coord in ("A2","A4"):
        ws[coord].fill = gray_fill
        ws[coord].font = bold
        ws[coord].alignment = left

    # 8) Bold & shade header row A6:F6
    for c in range(1, 7):
        cell = ws.cell(row=header_row, column=c)
        cell.font = bold_white
        cell.fill = dark_blue
    # 9) Shade A7 gray
    ws["A7"].fill = gray_fill

    # 10) Shade A9–A11, A13–A23, A25–A29, A31–A35 with tint_da
    for start,end in ((9,11),(13,23),(25,29),(31,35)):
        for r in range(start, end+1):
            ws.cell(row=r, column=1).fill = tint_da

    # 11) Shade rows 8,12,24,30 (A–F) light_blue
    for r in (8,12,24,30):
        for c in range(1,7):
            ws.cell(row=r, column=c).fill = light_blue

    # 12) Bold B7, B9,B10,B11,B13,B14,B17,B25,B29,B31,B32,B35
    for coord in ("B7","B9","B10","B11","B13","B14","B17","B25","B29","B31","B32","B35"):
        ws[coord].font = bold

    # 13) Maroon font in B15,B16,B18–B23,B26–B28,B33–B34
    for block in [(15,16),(18,23),(26,28),(33,34)]:
        for r in range(block[0], block[1]+1):
            ws.cell(row=r, column=2).font = maroon

    # 14) Fill C7:D7 with "-" and align C7:E7 right
    for c in range(3,4):
        ws.cell(row=7, column=c, value="-")
    for c in (3,4,5):
        ws.cell(row=5, column=c).alignment = right

    # 15) Column C formatting:
    fmt_map = {
        9: ("#,##0", None),    # C9 no decimals, thousand comma
        10: ("0.0", None),
        11: ("0.0", None),
        13: ("0.0", None),
    }
    for r, (nf, _) in fmt_map.items():
        ws.cell(row=r, column=3).number_format = nf
    # C18–C23 one decimal
    for r in range(18,24):
        ws.cell(row=r, column=3).number_format = "0.0"
    # C26–C29 one decimal
    for r in range(26,30):
        ws.cell(row=r, column=3).number_format = "0.0"
    # C31,C33,C34 one decimal
    for r in (31,33,34):
        ws.cell(row=r, column=3).number_format = "0.0"
    # C14,C17,C25,C32 replace with "-" and align right
    for r in (14,17,25,32):
        ws.cell(row=r, column=3, value="-").alignment = right
    # C15,C16,C35 no decimal
    for r in (15,16,35):
        c = ws.cell(row=r, column=3)
        if isinstance(c.value, (int, float)):
            c.number_format = "0"

    # 16) Column D 2 decimals for D9–D35
    for r in range(9,36):
        c = ws.cell(row=r, column=4)
        if isinstance(c.value, (int,float)):
            c.number_format = "0.00"

    # 17) Column E formatting & zero suppression
    for r in range(7,36):
        c = ws.cell(row=r, column=5)
        if isinstance(c.value, (int,float)):
            if c.value == 0:
                c.value = None
            else:
                c.number_format = "0.0"
    # bold+blue E9–E11, E13–E14, E17, E25,E29,E31,E32,E35
    for coord in ("E9","E10","E11","E13","E14","E17","E25","E29","E31","E32","E35"):
        c = ws[coord]; c.font = bold; c.font = Font(color="FF0000FF", bold=True)
    # maroon E15,E16,E18–E23,E26–E28,E33–E34
    for block in [(15,16),(18,23),(26,28),(33,34)]:
        for r in range(block[0], block[1]+1):
            ws.cell(row=r, column=5).font = maroon

    # 18) Maroon font in F15,F16,F18–F23,F26–F28,F33–F34
    for block in [(15,16),(18,23),(26,28),(33,34)]:
        for r in range(block[0], block[1]+1):
            ws.cell(row=r, column=6).font = maroon

    # 19) Column widths A–F
    widths = {"A":25,"B":52,"C":10,"D":12,"E":11.5,"F":57}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 20) Outside border A1:F35
    for r in range(1,36):
        for c in range(1,7):
            if r in (1,35) or c in (1,6):
                ws.cell(row=r, column=c).border = mk_border(
                    top   = (r==1),
                    bottom= (r==35),
                    left_b= (c==1),
                    right_b=(c==6)
                )

    # Save to buffer
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


excel_data_long = generate_custom_export_long(export_long_df)

# Put the Save + Export buttons side by side
# carve the page into 3 chunks: 
#  • 1 unit for btn1 
#  • 1 unit for btn2 
#  • 6 units of blank space

save_col_long, export_col_long, blank_col_long = st.columns([2, 2, 6])

with save_col_long:
    if st.button("💾 Save Analyst Overrides",key="long_save"):
        # Save only the override columns (factor-level edits) to a file
        columns_to_save_long = ["short_name", "Adjustment", "Analyst Comment"]
        updated_subset_long = updated_df_long[columns_to_save_long]

        # Use the full Google Sheet, then pass selected_name to target the right tab
        save_override_to_gsheet(sheet_long, updated_subset_long, selected_name, selected_year)

        # clear only the cache for fetch_overrides
        fetch_overrides_long.clear()

        st.success("✅ Overrides saved and rating updated.")
        st.rerun() #rerun entire script from top to bottom so analyst can see update immediately

with export_col_long:
    st.download_button(
    label="📥 Export to Excel (Formatted)",
    key="long_excel",
    data=excel_data_long,
    file_name="supp_rating_table.xlsx",
    mime="application/vnd.openxmlformats-officedocument-spreadsheetml.sheet"
)