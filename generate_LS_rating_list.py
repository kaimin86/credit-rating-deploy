import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from gsheets_utils import load_override_from_gsheet, save_override_to_gsheet
import os #--> helps to save user edits on to pc
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

import streamlit as st
import time

## set the references year that you want
choose_year = 2025

## get the relevant excel files in. Transform into df and dictionary where relevant

# get the name, rating, and predicted rating into a df
df_transform = pd.read_excel("transform_data.xlsx")
df_rating = df_transform.loc[
    df_transform['year'] == choose_year,
    ['name', 'rating', 'predicted_rating']
].reset_index(drop=True)

# get the ratings scale into an excel, and then into a dictonary
rating_index = pd.read_excel("index_rating_scale.xlsx")
#zip pairs the two columns row by row to help make into a dict
rating_dict = dict(zip(rating_index['Numeric'], rating_index['Credit Rating'])) 

# Extract list of unique countries
countries = df_rating['name'].unique().tolist()

## create a connection to google sheets

def init_gsheets_client():
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds_info = dict(st.secrets["gcp_service_account"])
    creds_info["private_key"] = creds_info["private_key"].replace("\\n", "\n")
    creds = Credentials.from_service_account_info(creds_info, scopes=scope)
    return gspread.authorize(creds)

client = init_gsheets_client()

sheet_short = client.open("analyst_overrides_short")

## pull out rating adjustments into a df

adjustment_records = []

for country in countries:
    try:
        print(f"⏳ Reading overrides for {country}…", end="", flush=True)
        ws = sheet_short.worksheet(country)
        records = ws.get_all_records()
        print(" done")                # shows you it finished
        time.sleep(1)
        df_sheet = pd.DataFrame(records)
        
        # ── DROP THE PREDICTED/FINAL ROWS ── 
        if "short_name" in df_sheet.columns:
            df_sheet = df_sheet.loc[
                ~df_sheet["short_name"].isin(["predicted_rating", "final_rating"])
            ]

        # If sheet is empty or no 'year' column, assume zero adjustment
        if 'year' not in df_sheet.columns or df_sheet.empty:
            total_adj = 0.0
        else:
            total_adj = (
                df_sheet.loc[df_sheet['year'] == choose_year, 'Adjustment']
                .sum()
            )
        adjustment_records.append({'name': country, 'Adjustment': total_adj})
    except gspread.exceptions.WorksheetNotFound:
        # If the tab is missing, treat adjustment as zero
        print(f"⚠️ {country} tab not found; assuming 0 adjustment")
        adjustment_records.append({'name': country, 'Adjustment': 0.0})
    time.sleep(1) # throttle before next iteration

df_adjustment = pd.DataFrame(adjustment_records)

## Merge main ratings df and the adjustment df
df_LS_rating = pd.merge(
    df_rating,
    df_adjustment,
    on='name',
    how='left'
)
df_LS_rating['Adjustment'] = df_LS_rating['Adjustment'].fillna(0.0)

#calculate LS_ratings
df_LS_rating['LS_rating'] = df_LS_rating['predicted_rating'] + df_LS_rating['Adjustment']
#force a min max to the LS_ratings
df_LS_rating['LS_rating'] = df_LS_rating['LS_rating'].clip(lower=1,upper=22)

#rename columns
df_LS_rating.rename(columns={
    "rating":"public_rating",
    "predicted_rating":"model_rating"
}, inplace=True)

#map LS_ratings to letter

def rating_to_letter(x):
    key = int(round(x))
    return rating_dict.get(key, "Unknown")

df_LS_rating["LS_letter"] = df_LS_rating["LS_rating"].apply(rating_to_letter)

#create dot columns to help shift within ERV

df_LS_rating["distance_lower_bound"] = df_LS_rating["LS_rating"] - (df_LS_rating["LS_rating"].round()-0.5) 
df_LS_rating.loc[df_LS_rating["LS_rating"] <= 1,  "distance_lower_bound"] = 0.0 #done to handle edge cases
df_LS_rating.loc[df_LS_rating["LS_rating"] >= 22, "distance_lower_bound"] = 1.0 #done to handle edge cases

# 1) helper to build a 21‑char line, dot moves from right (0.0) to left (1.0)
def build_dot_line(val, width=21):
    # compute position: 0-->20, 1-->0
    pos = round((1 - val) * (width - 1))
    # build blank line + dot
    line = " " * pos + "⚫" + " " * (width - 1 - pos)
    return line

# 2) apply to make a new column
df_LS_rating["ERV_Dot"] = (
    df_LS_rating["distance_lower_bound"]
    .round(2)                       # keep 2‑decimals if you like
    .apply(build_dot_line)
)

#import the coverage_list and merge into the main df
df_coverage = pd.read_excel("coverage_list.xlsx")
df_LS_rating = pd.merge(
    df_LS_rating,
    df_coverage[['name', 'in ERV', 'Analyst']],  # pick only the columns you need
    on="name",
    how="left"
)


## export fhe file into excel for sharing
output_path = "LS_rating.xlsx"

with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
    # 1) write the raw data
    df_LS_rating.to_excel(writer, index=False, sheet_name="Ratings")
    
    # 2) grab workbook & sheet
    wb = writer.book
    ws = writer.sheets["Ratings"]
    
    # 3) Auto‑fit widths on columns A, B, C, D, G and J
    for col in ["A", "B", "C", "D", "G","J"]:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in ws[col]
        )
        ws.column_dimensions[col].width = max_len + 2
    
    # 3.1) Force monospaced font + left‑align for the ERV_Dot column (H)
    for cell in ws["H"][1:]:   # skip header in row 1
        cell.font      = Font(name="Consolas")
        cell.alignment = Alignment(horizontal="left")
    
    # 3.2) force column H to width 26
    ws.column_dimensions['H'].width = 26

    # 4) Shade header row A1:J1 light‑blue
    header_fill = PatternFill("solid", fgColor="FFB6CEE4")
    for cell in ws[1]:
        cell.fill = header_fill

    # 5) Bold font for all cells A2:A138
    bold = Font(bold=True)
    for row in ws.iter_rows(min_row=2, max_row=138, min_col=1, max_col=1):
        for cell in row:
            cell.font = bold

    # 6) 2‑decimal formatting for B2:B138, C2:C138, D2:D138, E2:E138, G2:G138
    two_dp = "0.00"
    for col in ["B", "C", "D", "E", "G"]:
        for row in range(2, 139):
            ws[f"{col}{row}"].number_format = two_dp

    # 7) Apply thin border on every cell A1:J138
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=138, min_col=1, max_col=10):
        for cell in row:
            cell.border = border

    # 8) Turn on Excel’s AutoFilter for the block A1:J138
    ws.auto_filter.ref = "A1:J138"

# at this point writer.save() has been called
print(f"Exported and formatted LS Ratings to {output_path}")

#df_LS_rating.to_excel(output_path, index = False, engine="openpyxl")
#print(f"Exported LS Ratings to {output_path}")
